import streamlit as st
import qrcode
from PIL import Image
import io
import uuid
from datetime import datetime, timedelta
import random
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# Page config
st.set_page_config(
    page_title="Attendance System",
    page_icon="📋",
    layout="wide"
)

# Initialize session state
if 'active_sessions' not in st.session_state:
    st.session_state.active_sessions = {}
if 'student_codes' not in st.session_state:
    st.session_state.student_codes = {}
if 'current_session' not in st.session_state:
    st.session_state.current_session = None

# Courses database
COURSES = {
    "MS101-A": "Business Mathematics",
    "CS201-B": "Programming Fundamentals",
    "ENG101-C": "English Composition",
    "PHY101-D": "Physics for Engineers",
    "MTH101-E": "Calculus I"
}

def get_excel_path(course_code, course_title):
    """Get or create Excel file path"""
    filename = f"Attendance_Sheet_{course_code}_{course_title.replace(' ', '_')}.xlsx"
    filepath = os.path.join('attendance_sheets', filename)
    os.makedirs('attendance_sheets', exist_ok=True)
    
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.merge_cells('A1:C1')
        ws['A1'] = f"Attendance Sheet: {course_code} - {course_title}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A3'] = "Registration No"
        ws['B3'] = "Student Name"
        ws['C3'] = "Student Email"
        
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = Font(bold=True)
            ws[col].alignment = Alignment(horizontal='center')
        
        wb.save(filepath)
    
    return filepath

def mark_attendance(course_code, course_title, reg_no, date_str):
    """Mark attendance in Excel"""
    filepath = get_excel_path(course_code, course_title)
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Find or create date column
    date_col = None
    for col in range(4, ws.max_column + 2):
        if ws.cell(row=3, column=col).value == date_str:
            date_col = col
            break
        elif ws.cell(row=3, column=col).value is None:
            ws.cell(row=3, column=col, value=date_str)
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
            date_col = col
            break
    
    # Find or create student row
    student_row = None
    for row in range(4, ws.max_row + 2):
        if ws.cell(row=row, column=1).value == reg_no:
            student_row = row
            break
    
    if student_row is None:
        student_row = ws.max_row + 1
        ws.cell(row=student_row, column=1, value=reg_no)
        ws.cell(row=student_row, column=2, value="Unknown")
        ws.cell(row=student_row, column=3, value="unknown@example.com")
    
    # Mark attendance
    ws.cell(row=student_row, column=date_col, value="P")
    ws.cell(row=student_row, column=date_col).alignment = Alignment(horizontal='center')
    wb.save(filepath)
    return True

def cleanup_expired():
    """Remove expired sessions"""
    now = datetime.now()
    expired = []
    for sid, data in st.session_state.active_sessions.items():
        if now > datetime.fromisoformat(data['valid_until']):
            expired.append(sid)
    for sid in expired:
        del st.session_state.active_sessions[sid]
        if sid in st.session_state.student_codes:
            del st.session_state.student_codes[sid]
        if st.session_state.current_session == sid:
            st.session_state.current_session = None

# Main UI
st.title("📋 Smart Classroom Attendance System")
st.markdown("---")

# Sidebar
with st.sidebar:
    st.header("About")
    st.info("This system generates QR codes for class sessions. Students scan and enter their registration number to mark attendance.")
    st.markdown("---")
    st.markdown("**Made with:** Streamlit, QR Code, OpenPyXL")

# Tabs
tab1, tab2, tab3, tab4 = st.tabs(["👨‍🏫 Instructor", "👨‍🎓 Student", "📊 Records", "📈 Dashboard"])

# Tab 1: Instructor
with tab1:
    st.header("Generate QR Code")
    
    col1, col2 = st.columns(2)
    
    with col1:
        course_code = st.selectbox("Select Course", list(COURSES.keys()))
        course_title = COURSES[course_code]
        st.info(f"**Selected:** {course_code} - {course_title}")
        
        session_date = st.date_input("Class Date", datetime.now())
        valid_minutes = st.number_input("QR Validity (minutes)", min_value=5, max_value=120, value=15, step=5)
        
        if st.button("🚀 Generate QR Code", type="primary", use_container_width=True):
            session_id = str(uuid.uuid4())[:8]
            valid_until = datetime.now() + timedelta(minutes=valid_minutes)
            
            st.session_state.active_sessions[session_id] = {
                'course_code': course_code,
                'course_title': course_title,
                'date': session_date.strftime("%Y-%m-%d"),
                'valid_until': valid_until.isoformat()
            }
            st.session_state.student_codes[session_id] = {}
            st.session_state.current_session = session_id
            
            # Generate QR
            qr_data = f"Session: {session_id}\nCourse: {course_code}\nDate: {session_date}"
            qr = qrcode.QRCode(version=1, box_size=10, border=4)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_image = qr.make_image(fill_color="black", back_color="white")
            
            buf = io.BytesIO()
            qr_image.save(buf, format='PNG')
            st.image(buf, caption="Student QR Code", width=300)
            
            st.success(f"✅ Session Created!")
            st.info(f"**Session ID:** `{session_id}`\n**Valid until:** {valid_until.strftime('%H:%M:%S')}")
    
    with col2:
        st.subheader("Active Sessions")
        cleanup_expired()
        
        if st.session_state.active_sessions:
            for sid, data in st.session_state.active_sessions.items():
                with st.expander(f"Session: {sid[:6]}..."):
                    st.write(f"**Course:** {data['course_code']}")
                    st.write(f"**Date:** {data['date']}")
                    valid_until = datetime.fromisoformat(data['valid_until'])
                    st.write(f"**Expires:** {valid_until.strftime('%H:%M:%S')}")
                    st.write(f"**Students:** {len(st.session_state.student_codes.get(sid, {}))}")
                    if st.button(f"End", key=f"end_{sid}"):
                        del st.session_state.active_sessions[sid]
                        st.rerun()
        else:
            st.info("No active sessions")

# Tab 2: Student
with tab2:
    st.header("Mark Your Attendance")
    
    cleanup_expired()
    
    if not st.session_state.active_sessions:
        st.warning("⚠️ No active sessions. Please ask your instructor to generate a QR code.")
    else:
        session_options = {}
        for sid, data in st.session_state.active_sessions.items():
            valid_until = datetime.fromisoformat(data['valid_until'])
            if datetime.now() < valid_until:
                session_options[f"{data['course_code']} - {data['course_title']} ({data['date']})"] = sid
        
        if not session_options:
            st.warning("No active sessions available")
        else:
            selected = st.selectbox("Select Class", list(session_options.keys()))
            session_id = session_options[selected]
            session_data = st.session_state.active_sessions[session_id]
            
            st.info(f"**{session_data['course_code']}** - {session_data['course_title']}\n📅 {session_data['date']}")
            
            col1, col2 = st.columns(2)
            
            with col1:
                reg_no = st.text_input("Registration Number (20xxxxx)", max_chars=7, key="student_reg")
                
                if st.button("📱 Get 4-Digit Code", use_container_width=True):
                    if not reg_no or not reg_no.isdigit() or len(reg_no) != 7 or not reg_no.startswith('20'):
                        st.error("❌ Invalid registration number!")
                    else:
                        code = str(random.randint(1000, 9999))
                        code_expiry = datetime.now() + timedelta(minutes=2)
                        
                        if reg_no not in st.session_state.student_codes[session_id]:
                            st.session_state.student_codes[session_id][reg_no] = {}
                        
                        st.session_state.student_codes[session_id][reg_no]['code'] = code
                        st.session_state.student_codes[session_id][reg_no]['expiry'] = code_expiry.isoformat()
                        
                        st.success(f"🔢 Your code: **{code}**")
                        st.info(f"⏰ Expires at: {code_expiry.strftime('%H:%M:%S')}")
            
            with col2:
                verify_code = st.text_input("Enter 4-digit code", max_chars=4, key="student_code")
                
                if st.button("✅ Submit Attendance", type="primary", use_container_width=True):
                    if not reg_no:
                        st.error("Enter registration number first")
                    elif not verify_code:
                        st.error("Enter the 4-digit code")
                    elif reg_no not in st.session_state.student_codes.get(session_id, {}):
                        st.error("Generate a code first")
                    else:
                        code_data = st.session_state.student_codes[session_id][reg_no]
                        if datetime.now() > datetime.fromisoformat(code_data['expiry']):
                            st.error("Code expired! Get new code")
                        elif code_data['code'] != verify_code:
                            st.error("Invalid code!")
                        else:
                            mark_attendance(
                                session_data['course_code'],
                                session_data['course_title'],
                                reg_no,
                                session_data['date']
                            )
                            del st.session_state.student_codes[session_id][reg_no]
                            st.success(f"✅ Attendance marked for {reg_no}!")
                            st.balloons()

# Tab 3: Records
with tab3:
    st.header("View Attendance Records")
    
    course_view = st.selectbox("Select Course to View", list(COURSES.keys()), key="view_course")
    
    if st.button("📊 Load Attendance Sheet"):
        course_title = COURSES[course_view]
        filename = f"Attendance_Sheet_{course_view}_{course_title.replace(' ', '_')}.xlsx"
        filepath = os.path.join('attendance_sheets', filename)
        
        if os.path.exists(filepath):
            df = pd.read_excel(filepath, header=2)
            st.dataframe(df, use_container_width=True)
            
            # Download button
            with open(filepath, 'rb') as f:
                st.download_button("📥 Download Excel", f, file_name=filename)
        else:
            st.info("No attendance records found for this course")

# Tab 4: Dashboard
with tab4:
    st.header("System Dashboard")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Active Sessions", len(st.session_state.active_sessions))
    
    with col2:
        total_codes = sum(len(codes) for codes in st.session_state.student_codes.values())
        st.metric("Active Student Codes", total_codes)
    
    with col3:
        if os.path.exists('attendance_sheets'):
            files = len([f for f in os.listdir('attendance_sheets') if f.endswith('.xlsx')])
            st.metric("Total Sheets", files)
        else:
            st.metric("Total Sheets", 0)
    
    st.markdown("---")
    st.subheader("Recent Sessions")
    
    if st.session_state.active_sessions:
        for sid, data in list(st.session_state.active_sessions.items())[:5]:
            st.write(f"**{data['course_code']}** - {data['date']} (Expires: {datetime.fromisoformat(data['valid_until']).strftime('%H:%M')})")
    else:
        st.info("No active sessions")