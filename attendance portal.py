!pip install qrcode pillow openpyxl pandas

import qrcode
from PIL import Image
import ipywidgets as widgets
from IPython.display import display, clear_output, Image as IPImage
import uuid
from datetime import datetime, timedelta
import random
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import pandas as pd

# Global storage
active_sessions = {}
student_codes = {}
current_session = None

# Courses database
COURSES = {
    "MS101-A": "Business Mathematics",
    "CS201-B": "Programming Fundamentals", 
    "ENG101-C": "English Composition",
    "PHY101-D": "Physics for Engineers"
}

def get_excel_path(course_code, course_title):
    """Get or create Excel file path"""
    filename = f"Attendance_Sheet_{course_code}_{course_title.replace(' ', '_')}.xlsx"
    filepath = os.path.join('attendance_sheets', filename)
    os.makedirs('attendance_sheets', exist_ok=True)
    
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.merge_cells('A1:C1')
        ws['A1'] = f"Attendance Sheet: {course_code} - {course_title}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A3'] = "Registration No"
        ws['B3'] = "Student Name"
        ws['C3'] = "Student Email"
        wb.save(filepath)
    
    return filepath

def mark_attendance(course_code, course_title, reg_no, date_str):
    """Mark attendance in Excel"""
    filepath = get_excel_path(course_code, course_title)
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Find or create date column
    date_col = None
    for col in range(4, ws.max_column + 2):
        if ws.cell(row=3, column=col).value == date_str:
            date_col = col
            break
        elif ws.cell(row=3, column=col).value is None:
            ws.cell(row=3, column=col, value=date_str)
            ws.cell(row=3, column=col).font = Font(bold=True)
            date_col = col
            break
    
    # Find or create student row
    student_row = None
    for row in range(4, ws.max_row + 2):
        if ws.cell(row=row, column=1).value == reg_no:
            student_row = row
            break
    
    if student_row is None:
        student_row = ws.max_row + 1
        ws.cell(row=student_row, column=1, value=reg_no)
        ws.cell(row=student_row, column=2, value="Unknown")
        ws.cell(row=student_row, column=3, value="unknown@example.com")
    
    # Mark attendance
    ws.cell(row=student_row, column=date_col, value="P")
    wb.save(filepath)
    return True

def cleanup_expired():
    """Remove expired sessions"""
    now = datetime.now()
    expired = []
    for sid, data in active_sessions.items():
        if now > datetime.fromisoformat(data['valid_until']):
            expired.append(sid)
    for sid in expired:
        del active_sessions[sid]
        if sid in student_codes:
            del student_codes[sid]

print("✅ System ready!")

# Instructor Interface
print("="*60)
print("👨‍🏫 INSTRUCTOR PORTAL")
print("="*60)

# Create dropdown for courses
course_dropdown = widgets.Dropdown(
    options=[(f"{code} - {title}", code) for code, title in COURSES.items()],
    description='Course:',
    style={'description_width': 'initial'}
)

# Date picker
date_picker = widgets.DatePicker(
    description='Date:',
    value=datetime.now().date()
)

# Validity minutes
validity_slider = widgets.IntSlider(
    value=15,
    min=5,
    max=120,
    step=5,
    description='Valid (min):',
    style={'description_width': 'initial'}
)

# Generate button
generate_btn = widgets.Button(
    description='🚀 Generate QR Code',
    button_style='success',
    layout=widgets.Layout(width='200px')
)

# Output area
output_area = widgets.Output()

def generate_qr(b):
    global current_session
    with output_area:
        clear_output()
        
        course_code = course_dropdown.value
        course_title = COURSES[course_code]
        session_date = date_picker.value.strftime("%Y-%m-%d")
        valid_minutes = validity_slider.value
        
        # Create session
        session_id = str(uuid.uuid4())[:8]
        valid_until = datetime.now() + timedelta(minutes=valid_minutes)
        
        active_sessions[session_id] = {
            'course_code': course_code,
            'course_title': course_title,
            'date': session_date,
            'valid_until': valid_until.isoformat()
        }
        student_codes[session_id] = {}
        current_session = session_id
        
        # Generate QR code
        qr_data = f"Session: {session_id}\nCourse: {course_code}\nDate: {session_date}"
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Display QR
        display(qr_img)
        print(f"\n✅ SESSION CREATED!")
        print(f"📋 Session ID: {session_id}")
        print(f"⏰ Valid until: {valid_until.strftime('%H:%M:%S')}")
        print(f"📅 Date: {session_date}")
        print(f"📚 Course: {course_code} - {course_title}")
        
        # Save QR
        qr_img.save(f"qr_{session_id}.png")
        print(f"\n💾 QR saved as: qr_{session_id}.png")

generate_btn.on_click(generate_qr)

# Display instructor UI
display(course_dropdown, date_picker, validity_slider, generate_btn, output_area)

# View active sessions
view_sessions_btn = widgets.Button(description='📊 View Active Sessions', button_style='info')
sessions_output = widgets.Output()

def show_sessions(b):
    with sessions_output:
        clear_output()
        cleanup_expired()
        
        if not active_sessions:
            print("No active sessions")
            return
        
        print(f"Active Sessions ({len(active_sessions)}):\n")
        for sid, data in active_sessions.items():
            valid_until = datetime.fromisoformat(data['valid_until'])
            remaining = valid_until - datetime.now()
            print(f"🔵 Session: {sid}")
            print(f"   Course: {data['course_code']} - {data['course_title']}")
            print(f"   Date: {data['date']}")
            print(f"   Expires: {valid_until.strftime('%H:%M:%S')} ({remaining.seconds//60} min left)")
            print(f"   Students marked: {len(student_codes.get(sid, {}))}")
            print("-" * 40)

view_sessions_btn.on_click(show_sessions)
display(view_sessions_btn, sessions_output)

# Student Interface
print("\n" + "="*60)
print("👨‍🎓 STUDENT PORTAL")
print("="*60)

# Refresh button for sessions
refresh_btn = widgets.Button(description='🔄 Refresh Sessions', button_style='primary')
student_output = widgets.Output()

# Session selector (will be updated)
session_selector = widgets.Dropdown(description='Select Session:', style={'description_width': 'initial'})
reg_input = widgets.Text(description='Reg Number:', placeholder='20xxxxx', style={'description_width': 'initial'})
generate_code_btn = widgets.Button(description='🔢 Generate 4-Digit Code', button_style='warning')
code_input = widgets.Text(description='Enter Code:', placeholder='XXXX', style={'description_width': 'initial'})
submit_btn = widgets.Button(description='✅ Submit Attendance', button_style='success')
status_label = widgets.HTML(value="")

def update_sessions(b=None):
    """Update session dropdown"""
    cleanup_expired()
    options = []
    for sid, data in active_sessions.items():
        valid_until = datetime.fromisoformat(data['valid_until'])
        if datetime.now() < valid_until:
            label = f"{data['course_code']} - {data['course_title']} ({data['date']})"
            options.append((label, sid))
    
    session_selector.options = options
    if options:
        session_selector.value = options[0][1]
        return True
    else:
        return False

def generate_student_code(b):
    with student_output:
        clear_output(wait=True)
        
        if not session_selector.options:
            print("❌ No active sessions available")
            return
        
        reg_no = reg_input.value.strip()
        if not reg_no or not reg_no.isdigit() or len(reg_no) != 7 or not reg_no.startswith('20'):
            print("❌ Invalid registration number! Must be 7 digits starting with 20")
            return
        
        session_id = session_selector.value
        if session_id not in active_sessions:
            print("❌ Session expired")
            return
        
        # Generate 4-digit code
        code = str(random.randint(1000, 9999))
        code_expiry = datetime.now() + timedelta(minutes=2)
        
        if reg_no not in student_codes[session_id]:
            student_codes[session_id][reg_no] = {}
        
        student_codes[session_id][reg_no]['code'] = code
        student_codes[session_id][reg_no]['expiry'] = code_expiry.isoformat()
        
        print(f"✅ Your 4-digit code is: {code}")
        print(f"⏰ Expires at: {code_expiry.strftime('%H:%M:%S')}")
        print("\n⚠️ Enter this code below to confirm attendance")

def submit_attendance(b):
    with student_output:
        if not session_selector.options:
            print("❌ No active sessions")
            return
        
        reg_no = reg_input.value.strip()
        entered_code = code_input.value.strip()
        session_id = session_selector.value
        
        if session_id not in active_sessions:
            print("❌ Session expired")
            return
        
        if reg_no not in student_codes.get(session_id, {}):
            print("❌ Please generate a code first")
            return
        
        code_data = student_codes[session_id][reg_no]
        if datetime.now() > datetime.fromisoformat(code_data['expiry']):
            print("❌ Code expired! Generate new code")
            return
        
        if code_data['code'] != entered_code:
            print("❌ Invalid code!")
            return
        
        # Mark attendance
        session = active_sessions[session_id]
        mark_attendance(
            session['course_code'],
            session['course_title'],
            reg_no,
            session['date']
        )
        
        # Cleanup
        del student_codes[session_id][reg_no]
        print(f"✅ Attendance marked successfully for {reg_no}!")
        
        # Clear inputs
        code_input.value = ""

# Connect buttons
refresh_btn.on_click(update_sessions)
generate_code_btn.on_click(generate_student_code)
submit_btn.on_click(submit_attendance)

# Initial session update
update_sessions()

# Display student UI
display(refresh_btn, session_selector, reg_input, generate_code_btn, code_input, submit_btn, student_output)

# View attendance records
print("\n" + "="*60)
print("📊 VIEW ATTENDANCE RECORDS")
print("="*60)

course_view = widgets.Dropdown(
    options=[(f"{code} - {title}", code) for code, title in COURSES.items()],
    description='Course:'
)

view_btn = widgets.Button(description='📖 Show Attendance', button_style='info')
attendance_output = widgets.Output()

def show_attendance(b):
    with attendance_output:
        clear_output()
        
        course_code = course_view.value
        course_title = COURSES[course_code]
        
        filename = f"Attendance_Sheet_{course_code}_{course_title.replace(' ', '_')}.xlsx"
        filepath = os.path.join('attendance_sheets', filename)
        
        if os.path.exists(filepath):
            df = pd.read_excel(filepath, header=2)
            print(f"📊 Attendance Record: {course_code} - {course_title}\n")
            display(df)
            
            # Summary statistics
            print("\n📈 Summary:")
            for col in df.columns[3:]:  # Skip Reg No, Name, Email
                present = df[col].value_counts().get('P', 0)
                total = len(df)
                print(f"  {col}: {present}/{total} ({present*100/total:.1f}%)")
        else:
            print("No attendance records found for this course")

view_btn.on_click(show_attendance)
display(course_view, view_btn, attendance_output)

# Manual attendance entry (useful for testing)
print("\n" + "="*60)
print("🔧 MANUAL ATTENDANCE ENTRY (Testing)")
print("="*60)

manual_course = widgets.Dropdown(
    options=[(f"{code} - {title}", code) for code, title in COURSES.items()],
    description='Course:'
)

manual_reg = widgets.Text(description='Reg Number:', placeholder='20xxxxx')
manual_date = widgets.DatePicker(description='Date:', value=datetime.now().date())
manual_btn = widgets.Button(description='✏️ Mark Attendance', button_style='warning')
manual_output = widgets.Output()

def manual_mark(b):
    with manual_output:
        clear_output()
        
        course_code = manual_course.value
        course_title = COURSES[course_code]
        reg_no = manual_reg.value.strip()
        date_str = manual_date.value.strftime("%Y-%m-%d")
        
        if not reg_no or not reg_no.isdigit() or len(reg_no) != 7 or not reg_no.startswith('20'):
            print("❌ Invalid registration number")
            return
        
        mark_attendance(course_code, course_title, reg_no, date_str)
        print(f"✅ Attendance marked for {reg_no} on {date_str}")

manual_btn.on_click(manual_mark)
display(manual_course, manual_reg, manual_date, manual_btn, manual_output)

# System Dashboard
print("\n" + "="*60)
print("📊 SYSTEM DASHBOARD")
print("="*60)

dashboard_btn = widgets.Button(description='🔄 Refresh Dashboard', button_style='info')
dashboard_output = widgets.Output()

def show_dashboard(b):
    with dashboard_output:
        clear_output()
        
        cleanup_expired()
        
        print("📈 SYSTEM STATISTICS\n")
        print(f"Active Sessions: {len(active_sessions)}")
        print(f"Total Students with Codes: {sum(len(codes) for codes in student_codes.values())}")
        
        # List all attendance files
        if os.path.exists('attendance_sheets'):
            files = [f for f in os.listdir('attendance_sheets') if f.endswith('.xlsx')]
            print(f"\n📁 Attendance Sheets: {len(files)}")
            for f in files:
                filepath = os.path.join('attendance_sheets', f)
                df = pd.read_excel(filepath, header=2)
                total_students = len(df)
                print(f"  • {f}: {total_students} students")
        else:
            print("\nNo attendance sheets created yet")

dashboard_btn.on_click(show_dashboard)
display(dashboard_btn, dashboard_output)